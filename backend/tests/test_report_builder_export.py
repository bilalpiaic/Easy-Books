"""Report export (CSV/XLSX)."""
import io
from fastapi.testclient import TestClient


def _auth(client):
    client.post("/api/auth/signup", json={"email": "exp@rb.test", "password": "password123",
                                          "full_name": "U", "company_name": "Exp Co"})
    r = client.post("/api/auth/login", data={"username": "exp@rb.test", "password": "password123"})
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def _seed(client, auth):
    c = client.post("/api/customers", headers=auth, json={"name": "Acme"}).json()
    client.post("/api/invoices", headers=auth, json={
        "customer_id": c["id"], "issue_date": "2026-05-02", "due_date": "2026-06-30",
        "gst_rate": 0, "currency": "USD", "lines": [{"description": "S", "qty": 1, "rate": 1000}]})


def test_export_csv(client: TestClient):
    auth = _auth(client); _seed(client, auth)
    r = client.post("/api/report-builder/export?format=csv", headers=auth,
                    json={"source_key": "invoices", "config": {"columns": ["number", "total"]}})
    assert r.status_code == 200
    text = r.content.decode()
    assert "number,total" in text.replace(" ", "")
    assert "1000.00" in text


def test_export_xlsx_is_valid_workbook(client: TestClient):
    auth = _auth(client); _seed(client, auth)
    r = client.post("/api/report-builder/export?format=xlsx", headers=auth,
                    json={"source_key": "invoices", "config": {"columns": ["number", "total"]}})
    assert r.status_code == 200
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.active["A1"].value == "number"


def _seed_malicious(client, auth):
    c = client.post("/api/customers", headers=auth, json={"name": "=SUM(1+1)"}).json()
    client.post("/api/invoices", headers=auth, json={
        "customer_id": c["id"], "issue_date": "2026-05-02", "due_date": "2026-06-30",
        "gst_rate": 0, "currency": "USD", "lines": [{"description": "S", "qty": 1, "rate": 1000}]})


def test_export_csv_neutralises_formula_injection(client: TestClient):
    auth = _auth(client); _seed_malicious(client, auth)
    r = client.post("/api/report-builder/export?format=csv", headers=auth,
                    json={"source_key": "invoices", "config": {"columns": ["customer_name", "total"]}})
    assert r.status_code == 200
    text = r.content.decode()
    assert "'=SUM(1+1)" in text          # quoted -> inert
    assert "\n=SUM" not in text and not text.split("customer_name")[1].lstrip(",\r\n").startswith("=")


def test_export_xlsx_neutralises_formula_injection(client: TestClient):
    auth = _auth(client); _seed_malicious(client, auth)
    r = client.post("/api/report-builder/export?format=xlsx", headers=auth,
                    json={"source_key": "invoices", "config": {"columns": ["customer_name", "total"]}})
    assert r.status_code == 200
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.active["A2"].value == "'=SUM(1+1)"   # leading quote forces literal text
